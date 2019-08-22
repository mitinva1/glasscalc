#!/usr/bin/python3
# -*- coding: utf-8 -*-

import sys  # sys нужен для передачи argv в QApplication
import openpyxl
import datetime
import img
import os

from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *

import doorclass #doorclass
import door2  # desinger
import scene#this is scene

upGlass = doorclass.UpGlass('upGlass')
rightGlass = doorclass.UpGlass('upGlass')
leftGlass = doorclass.UpGlass('upGlass')
newDoor = doorclass.Door('New Door') #my door 1
newDoor.pushbase = 'DoorHandle'
#door places
botfitting = doorclass.BottomFitting('None')
botfitting2 = doorclass.BottomFitting('None')
doorglass = doorclass.Glass('None')
topfitting = doorclass.TopFitting('None')
topfitting2 = doorclass.TopFitting('None')
supportbf = doorclass.SupportBF('None')
supportbf2 = doorclass.SupportBF('None')
supporttf = doorclass.SupportTopFitting('None')
supporttf2 = doorclass.SupportTopFitting('None')
hinge = doorclass.Hinge('None'); hinge2 = doorclass.Hinge('None')
DoorHandle = doorclass.DoorHandle('None')
DoorHandle2 = doorclass.DoorHandle('None')
PushHandle = doorclass.PushHandle('None')
PushHandle2 = doorclass.PushHandle('None')
knob = doorclass.knob('None')
knob2 = doorclass.knob('None')
DoorLock = doorclass.DoorLock('None')
DoorLock2 = doorclass.DoorLock('None')
MateDoorLock = doorclass.MateDoorLock('None')
MateDoorLock2 = doorclass.MateDoorLock('None')
Compactor_down = doorclass.Compactor('None')
Compactor_down2 = doorclass.Compactor('None')
Compactor_side = doorclass.Compactor('None')
Compactor_side2 = doorclass.Compactor('None')
Compactor_up = doorclass.Compactor('None')
Compactor_up2 = doorclass.Compactor('None')
Compactor_hinge = doorclass.Compactor('None')
Compactor_hinge2 = doorclass.Compactor('None')
upProfile = doorclass.GlazingProfile('none')
leftProfile = doorclass.GlazingProfile('none')
rightProfile = doorclass.GlazingProfile('none')
downProfile = doorclass.GlazingProfile('none')
upCap = doorclass.GlazingCap('none')
leftCap = doorclass.GlazingCap('none')
rightCap = doorclass.GlazingCap('none')
downCap = doorclass.GlazingCap('none')


discount = 1.5
door_installation_price=0
inst_door_closer = 0



def connectbs(basename):#database for listwiew
        import sqlite3
        conn = sqlite3.connect("mydatabase.db")
        c = conn.cursor()
        c.execute('select * from %s order by code' %(basename))
        codelist = [x[0] for x in c]
        c.close()
        return codelist


def connectbs2(text, basename, xxx):#database for listwiew doorhandle
        import sqlite3
        conn = sqlite3.connect("mydatabase.db")
        c = conn.cursor()

        g = (text,); c.execute('select * from %s where code=?'%(basename), g)
        z2 = [x1 for x1 in c.fetchone()]
        if len(z2) == 5:
            xcx = xxx(z2[0],z2[1],z2[2],z2[3],z2[4])#xxx это сам класс
        else:
            xcx = xxx(z2[0],z2[1],z2[2],float(z2[3]))
        c.close()
        return xcx



class ExampleApp(QMainWindow, door2.Ui_GlassCalc, scene.myyScene):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.upt)
        self.pushButton_2.clicked.connect(self.screenshot)
#chekbox
        self.checkBox.stateChanged.connect(self.state_changed)
        self.checkBox_2.stateChanged.connect(self.state_changed_2)
        self.checkBox_3.stateChanged.connect(self.state_changed_3)#upGlass
        self.checkBox_4.stateChanged.connect(self.state_changed_4)#rightGlass
        self.checkBox_5.stateChanged.connect(self.state_changed_5)#leftGlass
#отключение глухих полей
        self.label_39.hide()
        self.label_36.hide()
        self.lineEdit_5.hide()
        self.lineEdit_6.hide()
        self.label_40.hide()
        self.label_37.hide()
        self.lineEdit_7.hide()
        self.lineEdit_8.hide()
        self.label_38.hide()
        self.label_41.hide()
        self.lineEdit_9.hide()
        self.lineEdit_10.hide()
#radiobuttn
        self.button_group_2 = QButtonGroup()
        self.button_group_2.addButton(self.radioButton_3)
        self.button_group_2.addButton(self.radioButton_4)
        self.button_group_2.buttonClicked.connect(self.on_radio)
        self.button_group = QButtonGroup()
        self.button_group.addButton(self.radioButton)
        self.button_group.addButton(self.radioButton_2)
        self.button_group.buttonClicked.connect(self.on_radio2)
        self.button_group_3 = QButtonGroup()
        self.button_group_3.addButton(self.radioButton_5)
        self.button_group_3.addButton(self.radioButton_6)
        self.button_group_3.addButton(self.radioButton_12)
        self.button_group_3.buttonClicked.connect(self.on_radio3)
        self.button_group_4 = QButtonGroup()
        self.button_group_4.addButton(self.radioButton_7)
        self.button_group_4.addButton(self.radioButton_8)
        self.button_group_4.buttonClicked.connect(self.on_radio4)
        self.button_group_5 = QButtonGroup()
        self.button_group_5.addButton(self.radioButton_9)
        self.button_group_5.addButton(self.radioButton_10)
        self.button_group_5.addButton(self.radioButton_11)
        self.button_group_5.buttonClicked.connect(self.on_radio5)
#comboBox
        self.comboBox_4.addItems(['без']+connectbs('glass'))
        self.comboBox_4.activated[str].connect(self.onActivated)
        self.comboBox_6.addItems(['без']+connectbs('BottomFitting'))
        self.comboBox_6.activated[str].connect(self.onActivated6)
        self.comboBox_7.addItems(['без']+connectbs('supportbf'))
        self.comboBox_7.activated[str].connect(self.onActivated7)
        self.comboBox_8.addItems(['без']+connectbs('TopFitting'))
        self.comboBox_8.activated[str].connect(self.onActivated8)
        self.comboBox_9.addItems(['без']+connectbs('supporttf'))
        self.comboBox_9.activated[str].connect(self.onActivated9)
        self.comboBox_10.addItems(['без']+connectbs('Hinge'))
        self.comboBox_10.activated[str].connect(self.onActivated10)
        self.comboBox_13.addItems(['без']+connectbs('BottomFitting'))
        self.comboBox_13.activated[str].connect(self.onActivated6_2)
        self.comboBox_14.addItems(['без']+connectbs('DoorLock'))
        self.comboBox_14.activated[str].connect(self.onActivated14)
        self.comboBox_15.addItems(['без']+connectbs('MateDoorLock'))
        self.comboBox_15.activated[str].connect(self.onActivated15)
        self.comboBox_16.addItems(['без']+connectbs('Compactor'))
        self.comboBox_16.activated[str].connect(self.onActivated16)
        self.comboBox_17.addItems(['без']+connectbs('Compactor'))
        self.comboBox_17.activated[str].connect(self.onActivated17)
        self.comboBox_18.addItems(['без']+connectbs('Compactor'))
        self.comboBox_18.activated[str].connect(self.onActivated18)
        self.comboBox_19.addItems(['без']+connectbs('supportbf'))
        self.comboBox_19.activated[str].connect(self.onActivated7_2)
        self.comboBox_20.addItems(['без']+connectbs('TopFitting'))
        self.comboBox_20.activated[str].connect(self.onActivated8_2)
        self.comboBox_21.addItems(['без']+connectbs('supporttf'))
        self.comboBox_21.activated[str].connect(self.onActivated9_2)
        self.comboBox_22.addItems(['без']+connectbs('Hinge'))
        self.comboBox_22.activated[str].connect(self.onActivated10_2)
        self.comboBox_25.addItems(['без']+connectbs('DoorLock'))
        self.comboBox_25.activated[str].connect(self.onActivated14_2)
        self.comboBox_26.addItems(['без']+connectbs('MateDoorLock'))
        self.comboBox_26.activated[str].connect(self.onActivated15_2)
        self.comboBox_27.addItems(['без']+connectbs('Compactor'))
        self.comboBox_27.activated[str].connect(self.onActivated27)
        self.comboBox_28.addItems(['без']+connectbs('Compactor'))
        self.comboBox_28.activated[str].connect(self.onActivated28)
        self.comboBox_29.addItems(['без']+connectbs('Compactor'))
        self.comboBox_29.activated[str].connect(self.onActivated29)
        self.comboBox_30.addItems(['без']+connectbs('Compactor'))
        self.comboBox_30.activated[str].connect(self.onActivated30)
        self.comboBox_31.addItems(['без']+connectbs('Compactor'))
        self.comboBox_31.activated[str].connect(self.onActivated31)
        self.comboBox_32.addItems(['Клиент', 'Дилер', '0%'])
        self.comboBox_32.activated[str].connect(self.onActivated32)
        self.comboBox_33.addItems(['верх без']+connectbs('profile40'))
        self.comboBox_33.activated[str].connect(self.onActivated33)
        self.comboBox_34.addItems(['лев без']+connectbs('profile40'))
        self.comboBox_34.activated[str].connect(self.onActivated34)
        self.comboBox_35.addItems(['прав. без']+connectbs('profile40'))
        self.comboBox_35.activated[str].connect(self.onActivated35)
        self.comboBox_36.addItems(['низ без']+connectbs('profile40'))
        self.comboBox_36.activated[str].connect(self.onActivated36)
        self.comboBox_37.addItems(['верх без']+connectbs('cap40'))
        self.comboBox_37.activated[str].connect(self.onActivated37)
        self.comboBox_38.addItems(['лев без']+connectbs('cap40'))
        self.comboBox_38.activated[str].connect(self.onActivated38)
        self.comboBox_39.addItems(['прав. без']+connectbs('cap40'))
        self.comboBox_39.activated[str].connect(self.onActivated39)
        self.comboBox_40.addItems(['низ без']+connectbs('cap40'))
        self.comboBox_40.activated[str].connect(self.onActivated40)
#scene
        self.der(*newDoor.for_scene(), *upGlass.for_scene(),
                 *rightGlass.for_scene(), *leftGlass.for_scene(),
                 *newDoor.for_scene2())
        self.lineEdit.setText('1000')
        self.lineEdit_2.setText('2000')
        self.lineEdit_3.setText('1000')
        self.lineEdit_4.setText('2000')
        self.update()

#radioButton
    def on_radio(self, button):#кол-во створок
        newDoor.quantitydoor = int(button.text())
        if newDoor.quantitydoor == 2:
            self.radioButton.hide()
            self.radioButton_2.hide()
            self.comboBox_13.show()
            self.comboBox_19.show()
            self.comboBox_20.show()
            self.comboBox_21.show()
            self.comboBox_22.show()
            self.comboBox_23.show()
            self.comboBox_25.show()
            self.comboBox_26.show()
            self.comboBox_27.show()
            self.comboBox_28.show()
            self.comboBox_29.show()
            self.comboBox_31.show()
        else:
            self.comboBox_13.hide()
            self.comboBox_19.hide()
            self.comboBox_20.hide()
            self.comboBox_21.hide()
            self.comboBox_22.hide()
            self.comboBox_23.hide()
            self.comboBox_25.hide()
            self.comboBox_26.hide()
            self.comboBox_27.hide()
            self.comboBox_28.hide()
            self.comboBox_29.hide()
            self.comboBox_31.hide()
            self.radioButton.show()
            self.radioButton_2.show()
#radioButton connect
    def on_radio2(self, button):#сторона открывания
        newDoor.sidedoor = button.text()

    def on_radio3(self, button):
        if button.text() == 'Офисная':
            self.comboBox_11.clear()
            self.comboBox_23.clear()
            self.label_11.setText('Офисная')
            newDoor.pushbase = 'PushHandle'
        elif button.text() == 'кноб':
            self.comboBox_11.clear()
            self.comboBox_23.clear()
            self.label_11.setText('кноб')
            newDoor.pushbase = 'knob'
        else:
            self.comboBox_11.clear()
            self.comboBox_23.clear()
            self.label_11.setText('нажимной гарнитур')
            newDoor.pushbase = 'DoorHandle'
            #self.comboBox_12.hide()
            #self.comboBox_11.show()
            #self.comboBox_24.hide()
        if button.text() != 'Офисная' and newDoor.quantitydoor==2:
            self.comboBox_23.clear()
            self.comboBox_23.show()
        if button.text() == 'Офисная' and newDoor.quantitydoor==2:
            self.comboBox_23.clear()
        self.comboBox_11.addItems(['без']+connectbs(newDoor.pushbase))
        self.comboBox_11.activated[str].connect(self.onActivated11)
        self.comboBox_23.addItems(['без']+connectbs(newDoor.pushbase))
        self.comboBox_23.activated[str].connect(self.onActivated11_2)

    def on_radio4(self, button):
        newDoor.hinge_quantity=int(button.text())

    def on_radio5(self, button):
        newDoor.doorlock_y2 = button.text()
        newDoor.doorlock_y2_2 = button.text()

#активаторы выпадающего списка
    def onActivated(self, text):#for glass
        global doorglass#denote the global variable
        xxx = doorclass.Glass#write to class variable for later transfer
        if text == 'без':
            doorglass = doorclass.Glass('None')
        else:
            basename = 'glass'
            doorglass = connectbs2(text, basename, xxx)#create an instance
    def onActivated6(self, text):#активация нижниго фитинга
        global botfitting
        xxx = doorclass.BottomFitting
        if text == 'без':
            botfitting = doorclass.BottomFitting('None')
        else:
            basename = 'BottomFitting'
            botfitting = connectbs2(text, basename, xxx)
    def onActivated6_2(self, text):#активация нижниго фитинга
        global botfitting2
        xxx = doorclass.BottomFitting
        if text == 'без':
            botfitting2 = doorclass.BottomFitting('None')
        else:
            basename = 'BottomFitting'
            botfitting2 = connectbs2(text, basename, xxx)
    def onActivated7(self, text):#activate support bf
        global supportbf
        xxx = doorclass.SupportBF
        if text == 'без':
            supportbf = doorclass.SupportBF('None')
        else:
            basename = 'supportbf'
            supportbf = connectbs2(text, basename, xxx)
    def onActivated7_2(self, text):#activate support bf
        global supportbf2
        xxx = doorclass.SupportBF
        if text == 'без':
            supportbf2 = doorclass.SupportBF('None')
        else:
            basename = 'supportbf'
            supportbf2 = connectbs2(text, basename, xxx)
    def onActivated8(self, text):#activate topfitting
        global topfitting
        xxx = doorclass.TopFitting
        if text == 'без':
            topfitting = doorclass.TopFitting('None')
        else:
            basename = 'TopFitting'
            topfitting = connectbs2(text, basename, xxx)
    def onActivated8_2(self, text):#activate topfitting
        global topfitting2
        xxx = doorclass.TopFitting
        if text == 'без':
            topfitting2 = doorclass.TopFitting('None')
        else:
            basename = 'TopFitting'
            topfitting2 = connectbs2(text, basename, xxx)
    def onActivated9(self, text):#активация нижниго фитинга
        global supporttf
        xxx = doorclass.SupportTopFitting
        if text == 'без':
            supporttf = doorclass.SupportTopFitting('None')
        else:
            basename = 'supporttf'
            supporttf = connectbs2(text, basename, xxx)
    def onActivated9_2(self, text):#активация нижниго фитинга
        global supporttf2
        xxx = doorclass.SupportTopFitting
        if text == 'без':
            supporttf2 = doorclass.SupportTopFitting('None')
        else:
            basename = 'supporttf'
            supporttf2 = connectbs2(text, basename, xxx)
    def onActivated10(self, text):#активация нижниго фитинга
        global hinge
        xxx = doorclass.Hinge
        if text == 'без':
            hinge = doorclass.Hinge('None')
        else:
            basename = 'Hinge'
            hinge = connectbs2(text, basename, xxx)
    def onActivated10_2(self, text):#активация нижниго фитинга
        global hinge2
        xxx = doorclass.Hinge
        if text == 'без':
            hinge2 = doorclass.Hinge('None')
        else:
            basename = 'Hinge'
            hinge2 = connectbs2(text, basename, xxx)
    def onActivated11(self, text):#активация нижниго фитинга
        global DoorHandle
        global PushHandle
        global knob
        if newDoor.pushbase == 'DoorHandle':
            PushHandle = doorclass.PushHandle('None')
            knob = doorclass.knob('none')
            xxx = doorclass.DoorHandle
            if text == 'без':
                DoorHandle = doorclass.DoorHandle('None')
            else:
                basename = 'DoorHandle'
                DoorHandle = connectbs2(text, basename, xxx)
        elif newDoor.pushbase == 'PushHandle':
            DoorHandle = doorclass.DoorHandle('None')
            knob = doorclass.knob('none')
            xxx = doorclass.PushHandle
            if text == 'без':
                PushHandle = doorclass.PushHandle('None')
            else:
                basename = 'PushHandle'
                PushHandle = connectbs2(text, basename, xxx)
        else:
            DoorHandle = doorclass.DoorHandle('None')
            PushHandle = doorclass.PushHandle('None')
            xxx = doorclass.knob
            if text == 'без':
                knob = doorclass.knob('None')
            else:
                basename = 'knob'
                knob = connectbs2(text, basename, xxx)
    def onActivated11_2(self, text):#активация нижниго фитинга
        global DoorHandle2
        global PushHandle2
        global knob2
        if newDoor.pushbase == 'DoorHandle':
            PushHandle2 = doorclass.PushHandle('None')
            knob2 = doorclass.knob('none')
            xxx = doorclass.DoorHandle
            if text == 'без':
                DoorHandle2 = doorclass.DoorHandle('None')
            else:
                basename = 'DoorHandle'
                DoorHandle2 = connectbs2(text, basename, xxx)
        elif newDoor.pushbase == 'PushHandle':
            knob2 = doorclass.knob('none')
            DoorHandle2 = doorclass.DoorHandle('None')
            xxx = doorclass.PushHandle
            if text == 'без':
                PushHandle2 = doorclass.PushHandle('None')
            else:
                basename = 'PushHandle'
                PushHandle2 = connectbs2(text, basename, xxx)
        else:
            DoorHandle2 = doorclass.DoorHandle('None')
            PushHandle2 = doorclass.PushHandle('None')
            xxx = doorclass.knob
            if text == 'без':
                knob2 = doorclass.knob('None')
            else:
                basename = 'knob'
                knob2 = connectbs2(text, basename, xxx)
    def onActivated14(self, text):#активация нижниго фитинга
        global DoorLock
        xxx = doorclass.DoorLock
        if text == 'без':
            DoorLock = doorclass.DoorLock('None')
        else:
            basename = 'DoorLock'
            DoorLock = connectbs2(text, basename, xxx)
    def onActivated14_2(self, text):#активация нижниго фитинга
        global DoorLock2
        xxx = doorclass.DoorLock
        if text == 'без':
            DoorLock2 = doorclass.DoorLock('None')
        else:
            basename = 'DoorLock'
            DoorLock2 = connectbs2(text, basename, xxx)
    def onActivated15(self, text):#активация нижниго фитинга
        global MateDoorLock
        xxx = doorclass.MateDoorLock
        if text == 'без':
            MateDoorLock=doorclass.MateDoorLock('None')
        else:
            basename = 'MateDoorLock'
            MateDoorLock = connectbs2(text, basename, xxx)
    def onActivated15_2(self, text):#активация нижниго фитинга
        global MateDoorLock2
        xxx = doorclass.MateDoorLock
        if text == 'без':
            MateDoorLock2=doorclass.MateDoorLock('None')
        else:
            basename = 'MateDoorLock'
            MateDoorLock2 = connectbs2(text, basename, xxx)
    def onActivated16(self, text):
        global Compactor_down
        xxx = doorclass.Compactor
        if text == 'без':
            Compactor_down=doorclass.Compactor('None')
        else:
            basename = 'Compactor'
            Compactor_down = connectbs2(text, basename, xxx)
    def onActivated17(self, text):
        global Compactor_side
        xxx = doorclass.Compactor
        if text == 'без':
            Compactor_side=doorclass.Compactor('None')
        else:
            basename = 'Compactor'
            Compactor_side = connectbs2(text, basename, xxx)
    def onActivated28(self, text):
        global Compactor_side2
        xxx = doorclass.Compactor
        if text == 'без':
            Compactor_side2=doorclass.Compactor('None')
        else:
            basename = 'Compactor'
            Compactor_side2 = connectbs2(text, basename, xxx)
    def onActivated18(self, text):
        global Compactor_up
        xxx = doorclass.Compactor
        if text == 'без':
            Compactor_up=doorclass.Compactor('None')
        else:
            basename = 'Compactor'
            Compactor_up = connectbs2(text, basename, xxx)
    def onActivated29(self, text):
        global Compactor_up2
        xxx = doorclass.Compactor
        if text == 'без':
            Compactor_up2=doorclass.Compactor('None')
        else:
            basename = 'Compactor'
            Compactor_up2 = connectbs2(text, basename, xxx)
    def onActivated30(self, text):
        global Compactor_hinge
        xxx = doorclass.Compactor
        if text == 'без':
            Compactor_hinge=doorclass.Compactor('None')
        else:
            basename = 'Compactor'
            Compactor_hinge = connectbs2(text, basename, xxx)
    def onActivated31(self, text):
        global Compactor_hinge2
        xxx = doorclass.Compactor
        if text == 'без':
            Compactor_hinge2 = doorclass.Compactor('None')
        else:
            basename = 'Compactor'
            Compactor_hinge2 = connectbs2(text, basename, xxx)
    def onActivated27(self, text):#активация нижниго фитинга
        global Compactor_down2
        xxx = doorclass.Compactor
        if text == 'без':
            Compactor_down2=doorclass.Compactor('None')
        else:
            basename = 'Compactor'
            Compactor_down2 = connectbs2(text, basename, xxx)
    def onActivated32(self, text):#активация дисконтной группы
        global discount
        xxx = doorclass.MateDoorLock
        if text == 'Клиент':
            discount = 1.5#наценка 50%
        elif text == 'Дилер':
            discount = 1.3#наценка 30%
        else:
            discount = 1#без наценки
    def onActivated33(self, text):#активация дисконтной группы
        global upProfile
        xxx = doorclass.GlazingProfile
        if text == 'верх без':
            upProfile = doorclass.GlazingProfile('none')
        else:
            basename = 'profile40'
            upProfile = connectbs2(text, basename, xxx)
    def onActivated34(self, text):#активация дисконтной группы
        global leftProfile
        xxx = doorclass.GlazingProfile
        if text == 'лев без':
            leftProfile = doorclass.GlazingProfile('none')
        else:
            basename = 'profile40'
            leftProfile = connectbs2(text, basename, xxx)
    def onActivated35(self, text):#активация дисконтной группы
        global rightProfile
        xxx = doorclass.GlazingProfile
        if text == 'прав. без':
            rightProfile = doorclass.GlazingProfile('none')
        else:
            basename = 'profile40'
            rightProfile = connectbs2(text, basename, xxx)
    def onActivated36(self, text):#активация дисконтной группы
        global downProfile
        xxx = doorclass.GlazingProfile
        if text == 'низ без':
            downProfile = doorclass.GlazingProfile('none')
        else:
            basename = 'profile40'
            downProfile = connectbs2(text, basename, xxx)
    def onActivated37(self, text):#активация дисконтной группы
        global upCap
        xxx = doorclass.GlazingProfile
        if text == 'верх без':
            upCap = doorclass.GlazingCap('none')
        else:
            basename = 'cap40'
            upCap = connectbs2(text, basename, xxx)
    def onActivated38(self, text):#активация дисконтной группы
        global leftCap
        xxx = doorclass.GlazingProfile
        if text == 'лев без':
            leftCap = doorclass.GlazingCap('none')
        else:
            basename = 'cap40'
            leftCap = connectbs2(text, basename, xxx)
    def onActivated39(self, text):#активация дисконтной группы
        global rightCap
        xxx = doorclass.GlazingProfile
        if text == 'прав. без':
            rightCap = doorclass.GlazingCap('none')
        else:
            basename = 'cap40'
            rightCap = connectbs2(text, basename, xxx)
    def onActivated40(self, text):#активация дисконтной группы
        global downCap
        xxx = doorclass.GlazingProfile
        if text == 'низ без':
            downCap = doorclass.GlazingCap('none')
        else:
            basename = 'cap40'
            downCap = connectbs2(text, basename, xxx)
#checkBox checkBox checkBox checkBox checkBox checkBox checkBox checkBox
    def state_changed(self, int):
        global door_installation_price
        if self.checkBox.isChecked():
            door_installation_price=1300
        else:
            door_installation_price=0

    def state_changed_2(self, int):
        global inst_door_closer
        if self.checkBox_2.isChecked():
            if newDoor.supportbf_name[2] == '8' or newDoor.supportbf_name[2] == '7':
                inst_door_closer =+ 1500
            if newDoor.supportbf_name2[2] == '8' or newDoor.supportbf_name[2] == '7':
                inst_door_closer =+ 1500
        else:
            inst_door_closer = 0
    def state_changed_3(self, int):
        if self.checkBox_3.isChecked():
            self.label_39.show()
            self.label_36.show()
            self.lineEdit_5.show()
            self.lineEdit_6.show()
            upGlass.name = doorglass.name
            upGlass.colour = doorglass.colour
        else:
            upGlass.name = 'none'
            upGlass.colour = 'none'
            upGlass.price = 0
            upGlass.width = 0
            upGlass.height = 0
            self.label_39.hide()
            self.label_36.hide()
            self.lineEdit_5.hide()
            self.lineEdit_6.hide()
            self.lineEdit_5.clear()
            self.lineEdit_6.clear()

    def state_changed_4(self, int):
        if self.checkBox_4.isChecked():
            self.label_40.show()
            self.label_37.show()
            self.lineEdit_7.show()
            self.lineEdit_8.show()
            rightGlass.name = doorglass.name
            rightGlass.colour = doorglass.colour
        else:
            rightGlass.name = 'none'
            rightGlass.colour = 'none'
            rightGlass.price = 0
            rightGlass.width = 0
            rightGlass.height = 0
            self.label_40.hide()
            self.label_37.hide()
            self.lineEdit_7.hide()
            self.lineEdit_8.hide()
            self.lineEdit_7.clear()
            self.lineEdit_8.clear()
    def state_changed_5(self, int):
        if self.checkBox_5.isChecked():
            self.label_41.show()
            self.label_38.show()
            self.lineEdit_9.show()
            self.lineEdit_10.show()
            leftGlass.name = doorglass.name
            leftGlass.colour = doorglass.colour
        else:
            leftGlass.name = 'none'
            leftGlass.colour = 'none'
            leftGlass.price = 0
            leftGlass.width = 0
            leftGlass.height = 0
            self.label_41.hide()
            self.label_38.hide()
            self.lineEdit_9.hide()
            self.lineEdit_10.hide()
            #self.lineEdit_9.clear()
            #self.lineEdit_10.clear()


#Вывод на печать
    def upt(self):
        self.update()

#filling class newDoor
        if len(self.lineEdit.text()) > 1:
            newDoor.doorx = float(self.lineEdit.text())
            newDoor.doory = float(self.lineEdit_2.text())
        else:
            pass
        if newDoor.quantitydoor == 2:
            newDoor.sidedoor = 'лев+прав'
            if len(self.lineEdit_3.text()) > 1:
                newDoor.doorx2 = float(self.lineEdit_3.text())
                newDoor.doory2 = float(self.lineEdit_4.text())
            else:
                pass
        else:
            newDoor.doorx2 = 0
            newDoor.doory2 = 0
        newDoor.glassname = doorglass.name
        if newDoor.quantitydoor == 1:
            newDoor.area = (newDoor.doory*newDoor.doorx/1000000)

        else:
            newDoor.area = (newDoor.doory*
                    newDoor.doorx/1000000)+(newDoor.doory2*
                    newDoor.doorx2/1000000)
#up Glass
        try:
            upGlass.width = float(self.lineEdit_5.text())
            upGlass.height = float(self.lineEdit_6.text())
        except:
            upGlass.width = 0
            upGlass.height = 0

        try:
            rightGlass.width = float(self.lineEdit_7.text())
            rightGlass.height = float(self.lineEdit_8.text())
        except:
            rightGlass.width = 0
            rightGlass.height = 0
        try:
            leftGlass.width = float(self.lineEdit_9.text())
            leftGlass.height = float(self.lineEdit_10.text())
        except:
            leftGlass.width = 0
            leftGlass.height = 0

        newDoor.doorglassprice = float(doorglass.price)*newDoor.area
        upGlass.price = doorglass.price
        rightGlass.price = doorglass.price
        leftGlass.price = doorglass.price
        newDoor.glass_colour = doorglass.colour
        newDoor.botfit_name = botfitting.name
        newDoor.botfit_name2 = botfitting2.name
        newDoor.botfit_colour = botfitting.colour
        newDoor.botfit_colour2 = botfitting2.colour
        newDoor.botfitprice = float(botfitting.price)
        newDoor.botfitprice2 = float(botfitting2.price)
        newDoor.supportbf_name = supportbf.name
        newDoor.supportbf_name2 = supportbf2.name
        newDoor.supportbf_colour = supportbf.colour
        newDoor.supportbf_colour2 = supportbf2.colour
        newDoor.supportbf_price = float(supportbf.price)
        newDoor.supportbf_price2 = float(supportbf2.price)
        newDoor.topfit_name = topfitting.name
        newDoor.topfit_colour = topfitting.colour
        newDoor.topfitprice = float(topfitting.price)
        newDoor.topfit_name2 = topfitting2.name
        newDoor.topfit_colour2 = topfitting2.colour
        newDoor.topfitprice2 = float(topfitting2.price)
        newDoor.supporttf_name = supporttf.name
        newDoor.supporttf_colour = supporttf.colour
        newDoor.supporttf_colour2 = supporttf2.colour
        newDoor.supporttf_price = float(supporttf.price)
        newDoor.supporttf_name2 = supporttf2.name
        newDoor.supporttf_price2 = float(supporttf2.price)
        newDoor.hinge_code = hinge.code
        newDoor.hinge_code2 = hinge2.code
        newDoor.hinge_name = hinge.name
        newDoor.hinge_price = float(hinge.price)
        newDoor.hinge_name2 = hinge2.name
        newDoor.hinge_price2 = float(hinge2.price)
        newDoor.doorhandle_code = DoorHandle.code
        newDoor.doorhandle_code2 = DoorHandle2.code
        newDoor.doorhandle_name = DoorHandle.name
        newDoor.doorhandle_price = float(DoorHandle.price)
        newDoor.doorhandle_name2 = DoorHandle2.name
        newDoor.doorhandle_price2 = float(DoorHandle2.price)
        newDoor.pushhandle_code = PushHandle.code
        newDoor.pushhandle_code2 = PushHandle2.code
        newDoor.pushhandle_name = PushHandle.name
        newDoor.pushhandle_l = int(PushHandle.length)
        newDoor.pushhandle_price = float(PushHandle.price)
        newDoor.pushhandle_name2 = PushHandle2.name
        newDoor.pushhandle_l2 = int(PushHandle2.length)
        newDoor.pushhandle_price2 = float(PushHandle2.price)
        newDoor.doorlock_code = DoorLock.code
        newDoor.doorlock_code2 = DoorLock2.code
        newDoor.doorlock_name = DoorLock.name
        newDoor.doorlock_price = float(DoorLock.price)
        newDoor.doorlock_name2 = DoorLock2.name
        newDoor.doorlock_price2 = float(DoorLock2.price)
        newDoor.matedoorlock_name = MateDoorLock.name
        newDoor.matedoorlock_price = float(MateDoorLock.price)
        newDoor.matedoorlock_name2 = MateDoorLock2.name
        newDoor.matedoorlock_price2 = float(MateDoorLock2.price)
        newDoor.compactor_down_name = Compactor_down.name
        newDoor.compactor_down_price = float(Compactor_down.price)
        newDoor.compactor_down_name2 = Compactor_down2.name
        newDoor.compactor_down_price2 = Compactor_down2.price
        newDoor.compactor_side_name = Compactor_side.name
        newDoor.compactor_side_price = Compactor_side.price
        newDoor.compactor_side_name2 = Compactor_side2.name
        newDoor.compactor_side_price2 = Compactor_side2.price
        newDoor.compactor_up_name = Compactor_up.name
        newDoor.compactor_up_price = Compactor_up.price
        newDoor.compactor_up_name2 = Compactor_up2.name
        newDoor.compactor_up_price2 = Compactor_up2.price
        newDoor.compactor_hinge_name = Compactor_hinge.name
        newDoor.compactor_hinge_price = Compactor_hinge.price
        newDoor.compactor_hinge_name2 = Compactor_hinge2.name
        newDoor.compactor_hinge_price2 = Compactor_hinge2.price
        newDoor.knob_code = knob.code
        newDoor.knob_code2 = knob2.code
        newDoor.knob_name = knob.name
        newDoor.knob_name2 = knob2.name
        newDoor.knob_price = float(knob.price)
        newDoor.knob_price2 = float(knob2.price)
        newDoor.left_profile_name = leftProfile.name
        newDoor.right_profile_name = rightProfile.name
        newDoor.leftCap_name = leftCap.name
        newDoor.rightCap_name = rightCap.name

        newDoor.upProfile_name = upProfile.name
        newDoor.upCap_name = upCap.name
        newDoor.downProfile_name = downProfile.name
        newDoor.downCap_name = downCap.name

#для верхнего профиля
        if (upGlass.width == 0 and rightGlass.width == 0 and leftGlass.width ==0):
            #upProfile = doorclass.GlazingProfile('none')

            self.comboBox_33.setCurrentIndex(0)
            self.comboBox_33.activated[str].connect(self.onActivated33)
            self.comboBox_34.setCurrentIndex(0)
            self.comboBox_35.setCurrentIndex(0)
            self.comboBox_36.setCurrentIndex(0)
            self.comboBox_37.setCurrentIndex(0)
            self.comboBox_38.setCurrentIndex(0)
            self.comboBox_39.setCurrentIndex(0)
            self.comboBox_40.setCurrentIndex(0)

            leftProfile.price = 0
            newDoor.left_profile_name = 'none'
            rightProfile.price = 0
            newDoor.right_profile_name = 'none'
            upProfile.price = 0
            newDoor.upProfile_name = 'none'
            upCap.price = 0
            newDoor.upCap_name = 'none'
            downProfile.price = 0
            newDoor.downProfile_name ='none'
            downCap.price = 0
            newDoor.downCap_name = 'none'

            leftCap.price = 0
            newDoor.leftCap_name = 'none'
            rightCap.price = 0
            newDoor.rightCap_name = 'none'

        elif (upGlass.width == 0 and rightGlass.width == 0):
            self.comboBox_35.setCurrentIndex(0)
            rightProfile.price = 0# = doorclass.GlazingProfile('none')
            newDoor.right_profile_name = 'none'
            rightCap.price = 0
            newDoor.rightCap_name = 'none'
            self.comboBox_39.setCurrentIndex(0)

        elif (upGlass.width == 0 and leftGlass.width ==0):
            self.comboBox_34.setCurrentIndex(0)
            leftProfile.price = 0#doorclass.GlazingProfile('none')
            newDoor.left_profile_name = 'none'
            leftCap.price = 0
            newDoor.leftCap_name = 'none'
            self.comboBox_38.setCurrentIndex(0)
            leftCap.price = 0#doorclass.GlazingCap('none')
        elif (leftGlass.width == 0 and  leftGlass.width ==0):
            self.comboBox_36.setCurrentIndex(0)
            self.comboBox_40.setCurrentIndex(0)

#downProfile
        try:
            newDoor.downProfile_length = (rightGlass.width + leftGlass.width)
            newDoor.downCap_length = (rightGlass.width + leftGlass.width)
        except:
            newDoor.downProfile_length = 0
            newDoor.downCap_length = 0
#upProfile
        try:
            newDoor.upProfile_length = (rightGlass.width + leftGlass.width +
                                       upGlass.width)
            newDoor.upCap_length = (rightGlass.width + leftGlass.width +
                                       upGlass.width)

        except:
            newDoor.upProfile_lengh = 0
            newDoor.upCap_length = 0
#left profile start

        if (leftGlass.height == 0 and upGlass.height != 0):
            try:
                newDoor.left_profile_length = upGlass.height
                newDoor.leftCap_length = upGlass.height
            except:
                newDoor.left_profile_length = 0
                newDoor.leftCap_length = 0
        else:
            newDoor.left_profile_length = leftGlass.height
            newDoor.leftCap_length = leftGlass.height

        if (rightGlass.height == 0 and upGlass.height != 0):
            try:
                newDoor.right_profile_length = upGlass.height
                newDoor.rightCap_length = upGlass.height
            except:
                newDoor.right_profile_length = 0
                newDoor.rightCap_length = 0
        else:
            newDoor.right_profile_length = rightGlass.height
            newDoor.rightCap_length = rightGlass.height

        newDoor.left_profile_price = (float(leftProfile.price)*
                                      newDoor.left_profile_length/1000)
        newDoor.right_profile_price = (float(rightProfile.price)*
                                      newDoor.right_profile_length/1000)
        newDoor.leftCap_price = (float(leftCap.price)*
                                 newDoor.leftCap_length/1000*2)
        newDoor.rightCap_price = (float(rightCap.price)*
                                  newDoor.rightCap_length/1000*2)
        newDoor.upProfile_price = (float(upProfile.price)*
                                  newDoor.upProfile_length/1000)
        newDoor.upCap_price = (float(upCap.price)*
                                  newDoor.upCap_length/1000*2)
        newDoor.downCap_price = (float(downCap.price)*
                                  newDoor.downCap_length/1000*2)
        newDoor.downProfile_price = (float(downProfile.price)*
                                  newDoor.downProfile_length/1000*2)

        if newDoor.left_profile_price == 0:
            newDoor.left_profile_length = 0
        if newDoor.right_profile_price == 0:
            newDoor.right_profile_length = 0
        if newDoor.leftCap_price == 0:
            newDoor.leftCap_length = 0
        if newDoor.rightCap_price == 0:
            newDoor.rightCap_length = 0
        upProfile.length_real = (upGlass.width + rightGlass.width +leftGlass.width)
#left profile end
        self.newScene()#call scene
        self.textBrowser.setText('готово')

        self.label_23.setText(str((newDoor.doorPrice() +
                             upGlass.priceup() + rightGlass.priceup() +
                             leftGlass.priceup())*
                             discount  +
                             (door_installation_price*
                             (newDoor.area + upGlass.area() +
                             rightGlass.area() + leftGlass.area()) +
                             inst_door_closer*newDoor.quantitydoor)))#цена двери на скидку + установка на площадь + уст. даводчиков на двери
#save image and exel file
    def screenshot(self):
        """
        save image in file and kp
        """
        date_kp = str(datetime.datetime.now()) #date our kp
        number_kp = date_kp[11:13]+date_kp[14:16]+date_kp[17:19] #number of kp
        image = QImage((newDoor.doorx + newDoor.doorx2 + leftGlass.width + rightGlass.width)/10, (upGlass.height+newDoor.doory)/10 + 10, QImage.Format_ARGB32_Premultiplied)
        painter = QPainter(image)
        # Render the region of interest to the QImage.
        self.scene.render(painter)
        painter.end()
        # Save the image to a file.
        if sys.platform == "linux" or sys.platform == "linux2":
            image.save("img/drw" + number_kp + ".png")
        else:
            image.save("img\\drw" + number_kp + ".png")

        wb = openpyxl.load_workbook(filename = 'kp3.xlsx')
        ws = wb.active
        if sys.platform == "linux" or sys.platform == "linux2":
            img = openpyxl.drawing.image.Image("img/drw" + number_kp + ".png")
        else:
            img = openpyxl.drawing.image.Image("img\\drw" + number_kp + ".png")
        ws.add_image(img, 'B11')
        rows=11

        ws['H9'] = number_kp
        ws['I10'] = date_kp[8:10] + '.' + date_kp[5:7] + '.' + date_kp[0:4]
        ws['F11'] = newDoor.accessories()

        if newDoor.quantitydoor == 2:
            ws['B36'] = 'Дверь цельно стеклянная двустворчатая'
        else:
            ws['B36'] = 'Дверь цельно стеклянная одностворчатая'
        ws['F36'] = '1'
        ws['H36'] = (newDoor.doorPrice() +
                    upGlass.priceup() + rightGlass.priceup() +
                    leftGlass.priceup())* discount
        ws['I36'] = (newDoor.doorPrice() +
                    upGlass.priceup() + rightGlass.priceup() +
                    leftGlass.priceup())* discount
        ws['F37'] = newDoor.area
        if door_installation_price != 0:
            ws['B37'] = 'монтаж'
            ws['H37'] = door_installation_price
            ws['I37'] = door_installation_price*(newDoor.area + upGlass.area() + rightGlass.area() + leftGlass.area())
        if inst_door_closer != 0:
            ws['B38'] = 'установка доводчика'
            ws['F38'] = newDoor.quantitydoor
            ws['G38'] = 'шт.'
            ws['H38'] = inst_door_closer
            ws['I38'] = inst_door_closer*newDoor.quantitydoor
        ws['I39'] = ((newDoor.doorPrice() +
                    upGlass.priceup() + rightGlass.priceup() +
                    leftGlass.priceup())*
                    discount  +
                    (door_installation_price*
                    (newDoor.area + upGlass.area() +
                    rightGlass.area() + leftGlass.area()) +
                    inst_door_closer*newDoor.quantitydoor))
        date_kp2 = date_kp[8:10] + '-' + date_kp[5:7] + '-' + date_kp[0:4]+'-' + number_kp

        if sys.platform == "linux" or sys.platform == "linux2":
            wb.save('newkp/kp'+date_kp2+'.xlsx')
            os.startfile(r'"newkp/kp'+date_kp2+'.xlsx"')
        else:
            wb.save('newkp\\kp'+date_kp2+'.xlsx')
            os.startfile(r'"newkp\\kp'+date_kp2+'.xlsx"')
        self.textBrowser.setText('Номер коммерческого ' + number_kp)
#my scene
    def newScene(self):
    #botFitting

        if botfitting.price != 0:
            if newDoor.sidedoor == 'правая':
                newDoor.botfitx = newDoor.doorx-161
            else:
                newDoor.botfitx = 0.01
        else:
            newDoor.botfitx = 0
        if botfitting2.price != 0 :
            newDoor.botfitx2 = newDoor.doorx+newDoor.doorx2+5-161
        else:
            newDoor.botfitx2 = 0
    #supportbf
        if supportbf.price != 0:
            if newDoor.supportbf_name[2] == '8' or newDoor.supportbf_name[2] == '7':
                newDoor.supportbf_w = 306
                newDoor.supportbf_l = 40
            else:
                newDoor.supportbf_w = 101
                newDoor.supportbf_l = 20
            if newDoor.sidedoor == 'правая':
                newDoor.supportbf_x = newDoor.doorx-newDoor.supportbf_w
            else:
                newDoor.supportbf_x = 0.01
        else:
            newDoor.supportbf_x = 0
        if supportbf2.price != 0:
            if newDoor.supportbf_name2[2] == '8' or newDoor.supportbf_name2[2] == '7':
                newDoor.supportbf_w2 = 306
                newDoor.supportbf_l2 = 40
            else:
                newDoor.supportbf_w2 = 101
                newDoor.supportbf_l2 = 20
            newDoor.supportbf_x2 = (newDoor.doorx + 5 + newDoor.doorx2 -
                                   newDoor.supportbf_w2)
        else:
            newDoor.supportbf_x2 = 0
    #topfitting
        if topfitting.price != 0:
            if newDoor.sidedoor == 'правая':
                newDoor.topfitx = newDoor.doorx - 161
            else:
                newDoor.topfitx = 0.01
        else:
            newDoor.topfitx = 0
        if topfitting2.price != 0:
            newDoor.topfitx2 = newDoor.doorx+newDoor.doorx2 + 5 - 161
        else:
            newDoor.topfitx2 = 0
    #supporttf
        if supporttf.price != 0:
            if newDoor.supporttf_name[2] == '8':
                newDoor.supporttf_w = 161
                newDoor.supporttf_l = 51
            else:
                newDoor.supporttf_w = 101
                newDoor.supporttf_l = 20
            if newDoor.sidedoor == 'правая':
                newDoor.supporttf_x = newDoor.doorx - newDoor.supporttf_w
            else:
                newDoor.supporttf_x = 0.01
        else:
            newDoor.supporttf_x = 0

        if supporttf2.price != 0:
            if newDoor.supporttf_name2[2] == '8':
                newDoor.supporttf_w2 = 161
                newDoor.supporttf_l2 = 51
            else:
                newDoor.supporttf_w2 = 101
                newDoor.supporttf_l2 = 20

            newDoor.supporttf_x2 = (newDoor.doorx+newDoor.doorx2 -
                                   newDoor.supporttf_w2)
        else:
            newDoor.supporttf_x2 = 0

    #hinge
        if hinge.price != 0:
            if newDoor.sidedoor == 'правая':
                newDoor.hinge_x = newDoor.doorx - 62 + 15
            else:
                newDoor.hinge_x = -15
        else:
            newDoor.hinge_x = 0
        if hinge2.price != 0:
            newDoor.hinge_x2 = newDoor.doorx+newDoor.doorx2 - 62 + 15
        else:
            newDoor.hinge_x2 = 0
    #doorhandle
        if DoorHandle.price != 0:
            if newDoor.sidedoor == 'левая' or newDoor.quantitydoor == 2:
                newDoor.doorhandle_x = newDoor.doorx - 120
                newDoor.doorhandle_x2 = newDoor.doorx - 170
            else:
                newDoor.doorhandle_x = 0.01
                newDoor.doorhandle_x2 = 0.01
        else:
            newDoor.doorhandle_x = 0

        if DoorHandle2.price != 0:
            newDoor.doorhandle_x_2 = newDoor.doorx + 5
            newDoor.doorhandle_x2_2 = newDoor.doorx + 5
        else:
            newDoor.doorhandle_x_2 = 0
#PushHandle
        if newDoor.pushhandle_price != 0:

            if newDoor.sidedoor == 'правая':
                newDoor.pushhandle_x = 68
            else:

                newDoor.pushhandle_x = newDoor.doorx - 132
        else:
            newDoor.pushhandle_x = 0

        if newDoor.pushhandle_price2 != 0:
            newDoor.pushhandle_x2 = newDoor.doorx + 68
        else:
            newDoor.pushhandle_x2 = 0
#DoorLock
        if DoorLock.price != 0:
            if newDoor.doorlock_y2 == 'с':
                newDoor.doorlock_w = 77
                newDoor.doorlock_l = 110
                newDoor.doorlock_y = newDoor.doory/2 - newDoor.doorlock_l/2
            elif newDoor.doorlock_y2 == 'в':
                newDoor.doorlock_w = 161
                newDoor.doorlock_l = 51
                newDoor.doorlock_y = 0
            else:
                newDoor.doorlock_w = 161
                newDoor.doorlock_l = 51
                newDoor.doorlock_y = newDoor.doory - newDoor.doorlock_l
            if newDoor.sidedoor == 'левая' or newDoor.quantitydoor == 2:
                newDoor.doorlock_x = newDoor.doorx - newDoor.doorlock_w
            else:
                newDoor.doorlock_x = 0.01
        else:
            newDoor.doorlock_x = 0

        if DoorLock2.price != 0:
            if newDoor.doorlock_y2_2 == 'с':
                newDoor.doorlock_w2 = 77
                newDoor.doorlock_l2 = 110
                newDoor.doorlock_y_2 = newDoor.doory2/2 - newDoor.doorlock_l2/2
            elif newDoor.doorlock_y2_2 == 'в':
                newDoor.doorlock_w2 = 161
                newDoor.doorlock_l2 = 51
                newDoor.doorlock_y_2 = 0
            else:
                newDoor.doorlock_w2 = 161
                newDoor.doorlock_l2 = 51
                newDoor.doorlock_y_2 = newDoor.doory2 - newDoor.doorlock_l2
            newDoor.doorlock_x2 = newDoor.doorx + 5
        else:
            newDoor.doorlock_x2 = 0

        self.der(*newDoor.for_scene(), *upGlass.for_scene(),
                 *rightGlass.for_scene(), *leftGlass.for_scene(),
                 *newDoor.for_scene2())
#scene screenshot

def main():
    app = QApplication(sys.argv)
    window = ExampleApp()
    window.show()
    app.exec_()

if __name__ == '__main__':
    main()
